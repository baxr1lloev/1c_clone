"""
Export utilities for generating PDF and Excel reports.
"""
from django.template.loader import render_to_string
from django.http import HttpResponse
import io


class ReportExporter:
    """Utility class for exporting reports to various formats."""
    
    @staticmethod
    def export_to_pdf(template_name, context, filename):
        """
        Generate PDF from HTML template using WeasyPrint.
        
        Args:
            template_name: Django template path
            context: Template context dictionary
            filename: Output PDF filename
            
        Returns:
            HttpResponse with PDF content
        """
        try:
            from weasyprint import HTML, CSS
        except ImportError:
            raise ImportError("WeasyPrint is not installed. Run: pip install weasyprint")
        
        # Render HTML from template
        html_string = render_to_string(template_name, context)
        
        # PDF-specific CSS styling
        css = CSS(string='''
            @page {
                size: A4;
                margin: 2cm;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 10pt;
            }
            h1 {
                font-size: 18pt;
                margin-bottom: 10pt;
            }
            h2 {
                font-size: 14pt;
                margin-top: 15pt;
                margin-bottom: 8pt;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-top: 10pt;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }
            th {
                background-color: #f3f4f6;
                font-weight: bold;
            }
            .text-right {
                text-align: right;
            }
            .total-row {
                font-weight: bold;
                background-color: #f9fafb;
            }
            .section-header {
                background-color: #e5e7eb;
                font-weight: bold;
                padding: 10px;
            }
        ''')
        
        # Generate PDF
        pdf_file = HTML(string=html_string).write_pdf(stylesheets=[css])
        
        # Create response
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def export_to_excel(workbook_data, filename):
        """
        Generate Excel file from structured data using openpyxl.
        
        Args:
            workbook_data: Dictionary with structure:
                {
                    'title': 'Report Title',
                    'sheets': [
                        {
                            'name': 'Sheet Name',
                            'headers': ['Col1', 'Col2', ...],
                            'rows': [[val1, val2, ...], ...]
                        }
                    ]
                }
            filename: Output Excel filename
            
        Returns:
            HttpResponse with Excel content
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("openpyxl is not installed. Run: pip install openpyxl")
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_data in workbook_data.get('sheets', []):
            ws = wb.create_sheet(title=sheet_data.get('name', 'Sheet1'))
            
            # Title
            if 'title' in sheet_data:
                ws['A1'] = sheet_data['title']
                ws['A1'].font = Font(size=16, bold=True)
                ws.merge_cells('A1:D1')
            
            # Headers
            headers = sheet_data.get('headers', [])
            header_row = 3 if 'title' in sheet_data else 1
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows
            start_row = header_row + 1
            rows = sheet_data.get('rows', [])
            
            for row_idx, row_data in enumerate(rows, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # Handle different data types
                    if isinstance(value, (int, float)):
                        cell.value = value
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.value = str(value)
                    
                    cell.border = border
                    
                    # Mark total rows
                    if isinstance(value, str) and 'Total' in value:
                        cell.fill = total_fill
                        cell.font = total_font
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
